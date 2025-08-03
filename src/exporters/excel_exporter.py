# src/exporters/excel_exporter.py
import pandas as pd
import json
import csv
from pathlib import Path
from typing import List, Dict, Any
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class TestCaseExporter:
    """Export test cases to Excel, CSV, and JSON formats"""
    
    def __init__(self):
        self.required_columns = [
            "User Story ID",
            "Acceptance Criteria ID", 
            "Scenario",
            "Test Case ID",
            "Test Case Description",
            "Precondition",
            "Steps",
            "Expected Result",
            "Part of Regression",
            "Priority"
        ]
    
    def export_to_excel(self, test_cases: List[Dict[str, Any]], output_path: str) -> str:
        """Export test cases to formatted Excel file"""
        try:
            # Create DataFrame
            df = pd.DataFrame(test_cases)
            
            # Ensure all required columns exist
            for col in self.required_columns:
                if col not in df.columns:
                    df[col] = ""
            
            # Reorder columns
            df = df[self.required_columns]
            
            # Create workbook with formatting
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"
            
            # Add headers
            for col_num, column_title in enumerate(self.required_columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Add data rows
            for row_num, test_case in enumerate(test_cases, 2):
                for col_num, column in enumerate(self.required_columns, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    value = test_case.get(column, "")
                    
                    # Handle multi-line content (Steps field)
                    if column == "Steps" and "\\n" in str(value):
                        value = value.replace("\\n", "\n")
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    cell.value = value
                    
                    # Apply conditional formatting based on Priority
                    if column == "Priority":
                        if value == "High":
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        elif value == "Medium":
                            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        elif value == "Low":
                            cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                    
                    # Add borders to all cells
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Auto-adjust column widths
            self._adjust_column_widths(ws)
            
            # Add summary sheet
            self._add_summary_sheet(wb, test_cases)
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Excel file exported successfully: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def export_to_csv(self, test_cases: List[Dict[str, Any]], output_path: str) -> str:
        """Export test cases to CSV format"""
        try:
            # Create DataFrame
            df = pd.DataFrame(test_cases)
            
            # Ensure all required columns exist
            for col in self.required_columns:
                if col not in df.columns:
                    df[col] = ""
            
            # Reorder columns
            df = df[self.required_columns]
            
            # Clean multi-line content for CSV
            for column in df.columns:
                df[column] = df[column].astype(str).str.replace('\\n', ' | ')
            
            # Export to CSV
            df.to_csv(output_path, index=False, encoding='utf-8')
            
            logger.info(f"CSV file exported successfully: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise
    
    def export_to_json(self, test_cases: List[Dict[str, Any]], output_path: str) -> str:
        """Export test cases to JSON format"""
        try:
            # Ensure all test cases have required fields
            cleaned_test_cases = []
            for case in test_cases:
                cleaned_case = {}
                for field in self.required_columns:
                    cleaned_case[field] = case.get(field, "")
                cleaned_test_cases.append(cleaned_case)
            
            # Export to JSON
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(cleaned_test_cases, f, indent=2, ensure_ascii=False)
            
            logger.info(f"JSON file exported successfully: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {str(e)}")
            raise
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths based on content"""
        column_widths = {
            'A': 15,  # User Story ID
            'B': 20,  # Acceptance Criteria ID
            'C': 25,  # Scenario
            'D': 15,  # Test Case ID
            'E': 40,  # Test Case Description
            'F': 30,  # Precondition
            'G': 50,  # Steps
            'H': 40,  # Expected Result
            'I': 18,  # Part of Regression
            'J': 12   # Priority
        }
        
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width
    
    def _add_summary_sheet(self, workbook, test_cases: List[Dict[str, Any]]):
        """Add a summary sheet with statistics"""
        summary_ws = workbook.create_sheet("Summary")
        
        # Calculate statistics
        total_cases = len(test_cases)
        priority_counts = {}
        regression_counts = {}
        user_story_counts = {}
        
        for case in test_cases:
            # Priority distribution
            priority = case.get("Priority", "Unknown")
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
            
            # Regression distribution
            regression = case.get("Part of Regression", "Unknown")
            regression_counts[regression] = regression_counts.get(regression, 0) + 1
            
            # User story distribution
            user_story = case.get("User Story ID", "Unknown")
            user_story_counts[user_story] = user_story_counts.get(user_story, 0) + 1
        
        # Add summary data
        summary_data = [
            ["Test Case Summary Report", ""],
            ["", ""],
            ["Total Test Cases", total_cases],
            ["", ""],
            ["Priority Distribution", ""],
            ["High Priority", priority_counts.get("High", 0)],
            ["Medium Priority", priority_counts.get("Medium", 0)],
            ["Low Priority", priority_counts.get("Low", 0)],
            ["", ""],
            ["Regression Test Distribution", ""],
            ["Part of Regression", regression_counts.get("Yes", 0)],
            ["Not in Regression", regression_counts.get("No", 0)],
            ["", ""],
            ["Coverage by User Story", ""],
        ]
        
        # Add user story coverage
        for story_id, count in user_story_counts.items():
            summary_data.append([story_id, count])
        
        # Write summary data to sheet
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_num, column=1, value=label)
            summary_ws.cell(row=row_num, column=2, value=value)
            
            # Format headers
            if "Distribution" in str(label) or "Summary Report" in str(label):
                summary_ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        
        # Adjust column widths
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15
    
    def export_all_formats(self, test_cases: List[Dict[str, Any]], base_filename: str) -> Dict[str, str]:
        """Export test cases to all supported formats"""
        results = {}
        
        try:
            base_path = Path(base_filename)
            
            # Excel export
            excel_path = base_path.with_suffix('.xlsx')
            results['excel'] = self.export_to_excel(test_cases, str(excel_path))
            
            # CSV export
            csv_path = base_path.with_suffix('.csv')
            results['csv'] = self.export_to_csv(test_cases, str(csv_path))
            
            # JSON export
            json_path = base_path.with_suffix('.json')
            results['json'] = self.export_to_json(test_cases, str(json_path))
            
            logger.info(f"All formats exported successfully to {base_path.parent}")
            return results
            
        except Exception as e:
            logger.error(f"Error exporting all formats: {str(e)}")
            raise
    
    def validate_test_cases(self, test_cases: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Validate test cases before export"""
        validation_report = {
            'total_cases': len(test_cases),
            'valid_cases': 0,
            'issues': []
        }
        
        for i, case in enumerate(test_cases):
            case_issues = []
            
            # Check required fields
            for field in self.required_columns:
                if not case.get(field, "").strip():
                    case_issues.append(f"Missing {field}")
            
            # Validate specific field formats
            if case.get("Priority") not in ["High", "Medium", "Low"]:
                case_issues.append("Invalid Priority value")
            
            if case.get("Part of Regression") not in ["Yes", "No"]:
                case_issues.append("Invalid Regression value")
            
            # Check minimum content length
            if len(case.get("Test Case Description", "")) < 10:
                case_issues.append("Test description too short")
            
            if len(case.get("Steps", "")) < 10:
                case_issues.append("Test steps too short")
            
            if case_issues:
                validation_report['issues'].append({
                    'case_index': i,
                    'test_case_id': case.get("Test Case ID", f"Case_{i}"),
                    'issues': case_issues
                })
            else:
                validation_report['valid_cases'] += 1
        
        validation_report['validation_percentage'] = (
            validation_report['valid_cases'] / validation_report['total_cases'] * 100
            if validation_report['total_cases'] > 0 else 0
        )
        
        return validation_report

# Usage example
if __name__ == "__main__":
    # Sample test cases
    sample_test_cases = [
        {
            "User Story ID": "US001",
            "Acceptance Criteria ID": "AC001",
            "Scenario": "Valid Login",
            "Test Case ID": "TC001",
            "Test Case Description": "Verify successful login with valid credentials",
            "Precondition": "User account exists and is active",
            "Steps": "1. Navigate to login page\\n2. Enter valid username\\n3. Enter valid password\\n4. Click Login button",
            "Expected Result": "User successfully logged in and redirected to dashboard",
            "Part of Regression": "Yes",
            "Priority": "High"
        }
    ]
    
    exporter = TestCaseExporter()
    
    # Validate test cases
    validation = exporter.validate_test_cases(sample_test_cases)
    print(f"Validation: {validation['valid_cases']}/{validation['total_cases']} cases valid")
    
    # Export to all formats
    results = exporter.export_all_formats(sample_test_cases, "test_cases")
    print(f"Exported files: {results}")